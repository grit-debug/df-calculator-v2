import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NAVY  = "002060"
WHITE = "FFFFFF"
BLACK = "000000"
MARJLAR = [
    (200, 1.80), (400, 1.60), (1000, 1.40), (2000, 1.20), (3000, 1.00),
    (4000, 0.90), (5000, 0.80), (10000, 0.75), (20000, 0.72),
    (50000, 0.69), (100000, 0.62),
]

FMT_USD = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_TRY = '_(₺* #,##0.00_);_(₺* (#,##0.00);_(₺* "-"??_);_(@_)'
FMT_NUM = '#,##0.00'

def parse_usd(s):
    try: return float(re.sub(r"[^\d.\-]", "", str(s))) or 0.0
    except: return 0.0

def navy_fill(): return PatternFill("solid", fgColor=NAVY)
def hdr_font(size=10): return Font(name="Arial", bold=True, size=size, color=WHITE)
def body_font(bold=False, size=10): return Font(name="Arial", bold=bold, size=size, color=BLACK)

def write_ozet(ws, snaps):
    ws.column_dimensions["A"].width = 20.875
    labels = ["Ürün Adı","Tarih","Mod","Baskı Tipi","USD/TRY Kuru",
              "Ürün Eni (cm)","Ürün Boyu (cm)","Sipariş Adedi",
              "Hammadde ($)","Baskı ($)","Üretim ($)","Toplam ($)"]
    for ri, label in enumerate(labels, 1):
        c = ws.cell(ri, 1, label)
        c.font = hdr_font()
        c.fill = navy_fill()
        ws.row_dimensions[ri].height = 18

    for ci, snap in enumerate(snaps, 2):
        bd = snap.get("baskiDetay") or {}
        vals = [
            snap.get("ad",""),
            snap.get("tarih",""),
            "Metre üzerinden" if snap.get("mode")=="metre" else "KG üzerinden",
            bd.get("tip","Yok"),
            snap.get("usdTry",""),
            snap.get("urunEn",""),
            snap.get("urunBoy",""),
            snap.get("sipAdet",""),
            parse_usd(snap.get("maliyetler",{}).get("kumas")),
            parse_usd(snap.get("maliyetler",{}).get("baski")),
            parse_usd(snap.get("maliyetler",{}).get("uretim")),
            parse_usd(snap.get("maliyetler",{}).get("toplam")),
        ]
        ws.column_dimensions[get_column_letter(ci)].width = 18.875
        for ri, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font = body_font(size=12)
            c.alignment = Alignment(horizontal="right")

def write_detay(ws, snap):
    ws.column_dimensions["A"].width = 30.875
    ws.column_dimensions["B"].width = 22.875
    ws.column_dimensions["C"].width = 18.875
    for i in range(1, 60):
        ws.row_dimensions[i].height = 18.75

    kur  = float(snap.get("usdTry") or 0)
    mode = snap.get("mode","metre")
    bd   = snap.get("baskiDetay") or None
    h    = snap.get("hammadde") or {}
    row  = [0]

    def wr(a, b=None, c=None, is_hdr=False, bold_b=False):
        row[0] += 1
        rn = row[0]
        ca = ws.cell(rn, 1, a)
        if is_hdr:
            ca.font = hdr_font()
            ca.fill = navy_fill()
            ca.alignment = Alignment(horizontal="center")
            ws.merge_cells(f"A{rn}:C{rn}")
        else:
            ca.font = body_font()
        if b is not None:
            cb = ws.cell(rn, 2, b)
            cb.font = body_font(bold=bold_b)
            cb.alignment = Alignment(horizontal="right") if isinstance(b,(int,float)) else Alignment()
            if isinstance(b, (int, float)):
                cb.number_format = FMT_USD
        if c is not None:
            cc = ws.cell(rn, 3, c)
            cc.font = body_font()
            cc.alignment = Alignment(horizontal="right")
            cc.number_format = FMT_TRY
        return rn

    def sec(t): wr(t, is_hdr=True)
    def bl():   row[0] += 1

    wr("DEMFABRIKA — Ürün Maliyet Kartı", is_hdr=True)
    wr("Ürün",         snap.get("ad",""))
    wr("Tarih",        snap.get("tarih",""))
    kur_row = wr("USD/TRY Kuru", kur)   # satır 4 = B4

    bl(); sec("── ÜRÜN BİLGİLERİ ──")
    wr("Ürün Eni (cm)",  snap.get("urunEn",""))
    wr("Ürün Boyu (cm)", snap.get("urunBoy",""))
    wr("Sipariş Adedi",  snap.get("sipAdet",""))

    bl(); sec("── HAMMADDE ──")
    wr("Ölçüm Modu", "Metre üzerinden" if mode=="metre" else "KG üzerinden")
    if mode == "metre":
        rr = wr("Metre Fiyatı", float(h.get("metreFiyat") or 0))
        ws.cell(rr,3, f"=B{rr}*B{kur_row}")
        wr("Kumaş Eni (cm)", float(h.get("kumasEn") or 0))
        wr("1m Şeride Ürün Adedi", float(h.get("urunAdet") or 0))
    else:
        rr = wr("KG Fiyatı", float(h.get("kgFiyat") or 0))
        ws.cell(rr,3, f"=B{rr}*B{kur_row}")
        wr("Kumaş Eni (cm)", float(h.get("kumasEn") or 0))
        wr("GSM (g/m²)",     float(h.get("gsm") or 0))
        wr("Metre/KG",       float(h.get("metreKg") or 0))
        wr("1m Şeride Ürün Adedi", float(h.get("urunAdet") or 0))
    hammadde_row = wr("Hammadde Maliyeti ($)",
                      parse_usd(snap.get("maliyetler",{}).get("kumas")), bold_b=True)
    ws.cell(hammadde_row,3, f"=B{hammadde_row}*B{kur_row}")

    bl(); sec("── BASKI ──")
    baski_usd = parse_usd(snap.get("maliyetler",{}).get("baski"))
    b5_row = None   # birim baskı +%5 satırı

    if not bd:
        wr("Baskı","Yok")
        b5_row = wr("Baskı Maliyeti (+%5 Pay)", 0, bold_b=True)
    elif bd.get("tip") == "Süblimasyon":
        wr("Baskı","Süblimasyon")
        rr = wr("Süblimasyon Fiyatı ($/m²)", float(bd.get("fiyat") or 0))
        ws.cell(rr,3, f"=B{rr}*B{kur_row}")
        wr("Baskı Adedi", float(bd.get("adet") or 0))
        b5_row = wr("Baskı Maliyeti (+%5 Pay)", baski_usd, bold_b=True)
        ws.cell(b5_row,3, f"=B{b5_row}*B{kur_row}")
    else:
        wr("Baskı","DTF")
        ref_en  = float(bd.get("refEn") or 0)
        ref_boy = float(bd.get("refBoy") or 0)
        ref_f   = float(bd.get("refFiyat") or 0)
        for i, bolge in enumerate(bd.get("bolgeler") or [], 1):
            wr(f"Baskı {i}- En",  float(bolge.get("en") or 0))
            wr(f"Baskı {i}- Boy", float(bolge.get("boy") or 0))

        ref_alan   = ref_en * ref_boy
        birim_alan = ref_f / ref_alan if ref_alan else 0
        dtf_only   = sum(birim_alan*(1+float(b.get("en",0))*float(b.get("boy",0)))
                         for b in (bd.get("bolgeler") or []))

        mesai = float(bd.get("pressMesai") or 0)
        maas  = float(bd.get("pressMaas") or 0)
        yrl   = float(bd.get("pressYerlestirme") or 0)
        sure  = float(bd.get("pressSure") or 0)
        n_bol = len(bd.get("bolgeler") or [])
        g_p   = mesai*60/(yrl+sure) if (yrl+sure)>0 else 0
        press_top = (maas/kur if kur else 0)/g_p*n_bol if g_p else 0

        rr = wr("Baskı Maliyeti", round(dtf_only,4))
        ws.cell(rr,3, f"=B{rr}*B{kur_row}")
        wr("Press - Yerleştirme Süresi (sn)", yrl)
        wr("Press - Pres Süresi (sn)", sure)
        rr = wr("Press Maliyeti", round(press_top,4))
        ws.cell(rr,3, f"=B{rr}*B{kur_row}")
        rr = wr("Birim Baskı Maliyeti", round(dtf_only+press_top,4))
        ws.cell(rr,3, f"=B{rr}*B{kur_row}")
        b5_row = wr("Birim Baskı Maliyeti (+%5 Pay)",
                    round((dtf_only+press_top)*1.05,4), bold_b=True)
        ws.cell(b5_row,3, f"=B{b5_row}*B{kur_row}")

    bl(); sec("── ÜRETİM MALİYETLERİ ──")
    row[0] += 1
    hr = row[0]
    for ci,t in [(1,"Kalem"),(2,"Tutar ($)"),(3,"Tutar (₺)")]:
        ws.cell(hr,ci,t).font = body_font()

    all_k = list(snap.get("sabitKalemler") or []) + list(snap.get("ekKalemler") or [])
    u_start = row[0]+1
    for kl in all_k:
        t  = float(kl.get("tutar") or 0)
        rr = wr(kl.get("ad",""), t)
        ws.cell(rr,3, f"=B{rr}*B{kur_row}")
    u_end = row[0]
    uretim_row = wr("Üretim Toplam ($)", f"=SUM(B{u_start}:B{u_end})", bold_b=True)
    ws.cell(uretim_row,2).number_format = FMT_USD
    ws.cell(uretim_row,3, f"=SUM(C{u_start}:C{u_end})")
    ws.cell(uretim_row,3).number_format = FMT_TRY

    bl(); sec("── MALİYET ÖZETİ ──")
    row[0] += 1
    hr2 = row[0]
    for ci,t in [(2,"Tutar ($)"),(3,"Tutar (₺)")]:
        ws.cell(hr2,ci,t).font = body_font()

    rr = wr("Hammadde", f"=B{hammadde_row}")
    ws.cell(rr,2).number_format = FMT_USD
    ws.cell(rr,3, f"=B{rr}*B{kur_row}"); ws.cell(rr,3).number_format = FMT_TRY
    rr = wr("Baskı", f"=B{b5_row}")
    ws.cell(rr,2).number_format = FMT_USD
    ws.cell(rr,3, f"=B{rr}*B{kur_row}"); ws.cell(rr,3).number_format = FMT_TRY
    rr = wr("Üretim", f"=B{uretim_row}")
    ws.cell(rr,2).number_format = FMT_USD
    ws.cell(rr,3, f"=B{rr}*B{kur_row}"); ws.cell(rr,3).number_format = FMT_TRY
    toplam_row = wr("BİRİM MALİYET TOPLAM",
                    f"=SUM(B{rr-2}:B{rr})", bold_b=True)
    ws.cell(toplam_row,2).number_format = FMT_USD
    ws.cell(toplam_row,3, f"=B{toplam_row}*B{kur_row}")
    ws.cell(toplam_row,3).number_format = FMT_TRY

    # C sütunundaki tüm hücrelere ₺ formatı uygula
    for i in range(1, row[0] + 1):
        c = ws.cell(i, 3)
        c.number_format = FMT_TRY

    return toplam_row, kur_row

def write_fiyat_listesi(ws, snap, toplam_row, kur_row, detay_name):
    ws.column_dimensions["A"].width = 17.0
    for col,w in [("B",14.0),("C",18.0),("D",18.0),("E",18.0),("F",18.0)]:
        ws.column_dimensions[col].width = w

    sip = int(snap.get("sipAdet") or 0)

    # Başlık satırı
    c = ws.cell(1,1,"FİYAT LİSTESİ")
    c.font = hdr_font(); c.fill = navy_fill()
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 18

    # Kolon başlıkları
    for ci,txt in enumerate(["Adet","Kar Marjı","Birim Satış (USD)",
                              "Birim Satış (TL)","Toplam (USD)","Toplam (TL)"],1):
        c = ws.cell(2,ci,txt)
        c.font = hdr_font(); c.fill = navy_fill()
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Sipariş adedine en yakın (>=) son kademedir
    highlight_idx = 0
    for i,(a,_) in enumerate(MARJLAR):
        if sip >= a: highlight_idx = i

    t_ref  = f"'{detay_name}'!B{toplam_row}"
    kr_ref = f"'{detay_name}'!B{kur_row}"

    for ri,(adet,marj) in enumerate(MARJLAR,3):
        bs_usd = f"={t_ref}*1.05*(1+{marj})"
        bs_tl  = f"={t_ref}*1.05*(1+{marj})*{kr_ref}"
        tot_usd= f"={t_ref}*1.05*(1+{marj})*{adet}"
        tot_tl = f"={t_ref}*1.05*(1+{marj})*{adet}*{kr_ref}"

        is_hl  = (ri-3) == highlight_idx and sip > 0
        for ci,val in enumerate([adet,f"{int(marj*100)}%",bs_usd,bs_tl,tot_usd,tot_tl],1):
            c = ws.cell(ri,ci,val)
            c.font = body_font(bold=is_hl, size=10)
            c.alignment = Alignment(horizontal="right" if ci>1 else "center")
            if ci in (3, 5): c.number_format = FMT_USD
            if ci in (4, 6): c.number_format = FMT_TRY
        ws.row_dimensions[ri].height = 15

def build_wb(snap: dict):
    wb = Workbook()
    ws_oz = wb.active; ws_oz.title = "Özet"
    safe = re.sub(r'[:\\/?*\[\]]','',str(snap.get("ad","Hesap")))[:28]
    ws_d = wb.create_sheet(safe)
    toplam_row, kur_row = write_detay(ws_d, snap)
    ws_fl = wb.create_sheet("Fiyat Listesi")
    write_fiyat_listesi(ws_fl, snap, toplam_row, kur_row, safe)
    wb.remove(ws_oz)
    return wb

def build_wb_multi(snaps: list):
    wb = Workbook()
    ws_oz = wb.active; ws_oz.title = "_tmp"
    for snap in snaps:
        safe = re.sub(r'[:\\/?*\[\]]','',str(snap.get("ad","Hesap")))[:28]
        ws_d = wb.create_sheet(safe)
        toplam_row, kur_row = write_detay(ws_d, snap)
        ws_fl = wb.create_sheet(safe[:24]+"_FL")
        write_fiyat_listesi(ws_fl, snap, toplam_row, kur_row, safe)
    wb.remove(wb["_tmp"])
    return wb
